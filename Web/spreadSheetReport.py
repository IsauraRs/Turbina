from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import (ScatterChart, Reference, Series,)

headersL = []
dg = []
tdg = []
row1 = []
row2 = []
row3 = []
row4 = []

filesheet = "ReporteSpreadsheet.xlsx"
wb = load_workbook(filesheet)
graph = wb.create_sheet(index = 2, title = "RvsRPM")
graph2 = wb.create_sheet(index = 3, title = "RvsEfm")
graph3 = wb.create_sheet(index = 4, title = "RvsEft")
#datos = obtener_etiquetas(an2.ciudades,an2.edad_por_grupos,an2.sexo,an2.nse, an2.marcas)
sheet = wb.active


logos = Image("static/img/logoSS.png")
sheet.add_image(logos,'A1')

sheet.delete_rows(6,7920)

def getHeaders():
    A1 = sheet['A5'].value
    B1 = sheet['B5'].value
    C1 = sheet['C5'].value
    D1 = sheet['D5'].value
    E1 = sheet['E5'].value
    F1 = sheet['F5'].value
    G1 = sheet['G5'].value
    H1 = sheet['H5'].value
    I1 = sheet['I5'].value
    J1 = sheet['J5'].value

    headersL = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1']

    s = [headersL,]
    #print("S antes: ", s)
def setData(data):
    #print("dataa:",data)
    for i in data:

        sheet.append(i)
        #print("sh",sheet)

    wb.save(filesheet)

def createGraphs(disl,disl1): #,disl2,disl3):
    #for i in range(1,len(disl)):

    res = 'Resistencia'
    rps = 'RPM'
    disl = [res] + disl
    disl1 = [rps] + disl1


    for g in range(1): 
        graph.append(disl)
        graph.append(disl1)
    
    chart = ScatterChart()
    chart.title = 'R[Ω] vs RPM'
    chart.x_axis.title = 'RPM'
    chart.y_axis.title = 'R[Ω]'
    xvals = Reference(graph, min_col=2,min_row=1,max_col=len(disl))
    #for s in range(1,len(disl)): #10
    values = Reference(graph,min_col=2,min_row=2,max_col=len(disl))
    series = Series(values,xvals,title_from_data=True)
    chart.series.append(series)
    #chart.add_data(values)
    #s = chart.series[1]
    graph.add_chart(chart,"M1")
    wb.save(filesheet)

def createGraph2(disl,disl2): #,disl2,disl3):
    
    res = 'Resistencia'
    efg = 'Ef.Generador'
    disl = [res] + disl
    disl2 = [efg] + disl2

    for g in range(1): 
        graph2.append(disl)
        graph2.append(disl2)
        
    
    chart = ScatterChart()
    chart.title = 'R[Ω] vs Eficiencia Generador'
    chart.x_axis.title = 'Efm'
    chart.y_axis.title = 'R[Ω]'
    xvals = Reference(graph2, min_col=2,min_row=1,max_col=len(disl))
    #for s in range(1,len(disl)): #10
    values = Reference(graph2,min_col=2,min_row=2,max_col=len(disl))
    series = Series(values,xvals,title_from_data=True)
    chart.series.append(series)
    #chart.add_data(values)
    #s = chart.series[1]
    graph2.add_chart(chart,"M1")
    wb.save(filesheet)

def createGraph3(disl,disl3): #,disl2,disl3):
    
    res = 'Resistencia'
    eft = 'Ef.Turbina'
    disl = [res] + disl
    disl3 = [eft] + disl3


    #for i in range(1,len(disl)):

    for g in range(1): 
        graph3.append(disl)
        graph3.append(disl3)
        
    
    chart = ScatterChart()
    chart.title = 'R[Ω] vs Eficiencia turbina'
    chart.x_axis.title = 'Eft'
    chart.y_axis.title = 'R'
    xvals = Reference(graph3, min_col=2,min_row=1,max_col=len(disl))
    #for s in range(1,len(disl)): #10
    values = Reference(graph3,min_col=2,min_row=2,max_col=len(disl))
    series = Series(values,xvals,title_from_data=True)
    chart.series.append(series)
    #chart.add_data(values)
    #s = chart.series[1]
    graph3.add_chart(chart,"M1")
    wb.save(filesheet)

